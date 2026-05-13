"""
Bulk upload views for Properties (CSV / Excel).

POST /api/v1/properties/bulk-upload/
    Accepts a multipart file (CSV or .xlsx).
    Returns a summary: created, skipped (duplicate name), errors per row.

GET  /api/v1/properties/bulk-upload/template/?format=csv|xlsx
    Returns a downloadable template with sample data.
"""

import csv
import io

from django.http import HttpResponse
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import IsWorkspaceAdminOrReadOnly
from properties.models import Property, PropertyLocation

# ---------------------------------------------------------------------------
# Column definitions — single source of truth for template + docs
# ---------------------------------------------------------------------------

PROPERTY_COLUMNS = [
    # (header, required, description, valid_values_or_None, example)
    ("name",              True,  "Unique property name",                                               None,                           "Sunset Gardens Phase 2"),
    ("property_type",     False, "Type of property. Defaults to RESIDENTIAL_LAND if blank.",           "RESIDENTIAL_LAND|FARM_LAND|COMMERCIAL|MIXED_USE", "RESIDENTIAL_LAND"),
    ("description",       False, "Free-text description of the property",                              None,                           "Premium residential land with good road network and drainage"),
    ("estate_land_title", False, "Title deed / C of O reference number",                               None,                           "C of O No. LAG/2024/12345"),
    ("status",            False, "Publication status. Defaults to DRAFT if blank.",                    "DRAFT|PUBLISHED",              "DRAFT"),
    ("city",              True,  "City where the property is located",                                 None,                           "Ibeju-Lekki"),
    ("state",             True,  "State where the property is located",                                None,                           "Lagos"),
    ("address",           False, "Street address (auto-generated from city+state if blank)",           None,                           "KM 35 Lekki-Epe Expressway"),
    ("postal_code",       False, "Postal / ZIP code",                                                  None,                           "106104"),
    ("nearest_landmark",  False, "Notable landmark near the property",                                 None,                           "Opposite La Campagne Resort"),
    ("latitude",          False, "GPS latitude in decimal degrees (e.g. 6.524379)",                   None,                           "6.4523"),
    ("longitude",         False, "GPS longitude in decimal degrees (e.g. 3.379206)",                  None,                           "3.8741"),
]

PROPERTY_TEMPLATE_HEADERS = [c[0] for c in PROPERTY_COLUMNS]

# Multiple sample rows covering different property types / scenarios
PROPERTY_SAMPLE_ROWS = [
    ["Sunset Gardens Phase 1",   "RESIDENTIAL_LAND", "Premium residential plots in a fully serviced estate with perimeter fencing, road network and estate gate.", "C of O No. LAG/2024/001", "PUBLISHED", "Ibeju-Lekki",  "Lagos",    "KM 35 Lekki-Epe Expressway", "106104", "Near Eleko Beach", "6.4523", "3.8741"],
    ["Green Valley Farm Estate", "FARM_LAND",        "Arable farmland suitable for crop cultivation and agro-processing.",                                        "Survey Plan No. OG/789",  "DRAFT",     "Abeokuta",    "Ogun",     "Sagamu-Abeokuta Road",       "",       "Near FUNAAB",      "7.1475", "3.3619"],
    ["Victoria Commerce Hub",    "COMMERCIAL",       "Commercial plots in a high-traffic business district with 24/7 power and fibre connectivity.",               "",                        "DRAFT",     "Victoria Island","Lagos", "Adeola Odeku Street",        "101241", "Near Four Points Hotel", "6.4281", "3.4219"],
    ["Harmony Mixed-Use Park",   "MIXED_USE",        "Mixed residential and commercial plots ideal for buy-to-let investment.",                                    "Gov't Consent No. KN/44", "DRAFT",     "Kano",         "Kano",    "Audu Bako Way",              "",       "Near Kano State Secretariat", "12.0022", "8.5919"],
    ["Palm Court Residences",    "RESIDENTIAL_LAND", "Gated estate with security, electricity and tarred internal roads.",                                         "C of O No. ABO/2023/55",  "PUBLISHED", "Kubwa",        "Abuja",   "Airport Road",               "900108", "Off Nnamdi Azikiwe International Airport", "9.0082", "7.3917"],
]

PROPERTY_TYPE_MAP = {
    "residential land": "RESIDENTIAL_LAND",
    "residential_land": "RESIDENTIAL_LAND",
    "farm land": "FARM_LAND",
    "farm_land": "FARM_LAND",
    "farmland": "FARM_LAND",
    "commercial": "COMMERCIAL",
    "commercial land": "COMMERCIAL",
    "mixed use": "MIXED_USE",
    "mixed_use": "MIXED_USE",
    "mixeduse": "MIXED_USE",
}
VALID_PROPERTY_TYPES = set(PROPERTY_TYPE_MAP.values())

STATUS_MAP = {
    "draft": "DRAFT",
    "published": "PUBLISHED",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_rows(file_obj, filename: str):
    """Return list of row-dicts from CSV or xlsx."""
    name = filename.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl is required for Excel uploads. Use CSV instead.")
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        # Use the first sheet named "Properties" if present, else the active sheet
        ws = wb["Properties"] if "Properties" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        data_rows = []
        for row in rows[1:]:
            d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            # Skip rows that are entirely empty or are the description row
            if all(not v for v in d.values()):
                continue
            data_rows.append(d)
        return data_rows
    else:
        text = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            d = {k.strip().lower(): (v.strip() if v else "") for k, v in row.items()}
            # Skip rows starting with "#" (comment/description rows)
            if d.get(PROPERTY_TEMPLATE_HEADERS[0], "").startswith("#"):
                continue
            rows.append(d)
        return rows


# ---------------------------------------------------------------------------
# Upload view
# ---------------------------------------------------------------------------

class PropertyBulkUploadView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated, IsWorkspaceAdminOrReadOnly]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            rows = _parse_rows(file, file.name)
        except Exception as e:
            return Response({"detail": str(e)}, status=400)

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=400)

        if len(rows) > 500:
            return Response({"detail": "Maximum 500 rows per upload. Please split your file."}, status=400)

        created = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(rows, start=2):
            name = row.get("name", "").strip()
            if not name:
                errors.append({"row": idx, "error": "name is required"})
                continue

            # property_type
            raw_type = row.get("property_type", "").strip().lower()
            prop_type = PROPERTY_TYPE_MAP.get(raw_type) or (raw_type.upper() if raw_type.upper() in VALID_PROPERTY_TYPES else "RESIDENTIAL_LAND")

            # status
            raw_status = row.get("status", "").strip().lower()
            status = STATUS_MAP.get(raw_status, "DRAFT")

            city = row.get("city", "").strip()
            state = row.get("state", "").strip()
            if not city:
                errors.append({"row": idx, "error": f"city is required (name='{name}')"})
                continue
            if not state:
                errors.append({"row": idx, "error": f"state is required (name='{name}')"})
                continue

            if Property.objects.filter(workspace=request.workspace, name__iexact=name).exists():
                skipped += 1
                continue

            try:
                lat_raw = row.get("latitude", "").strip()
                lng_raw = row.get("longitude", "").strip()
                lat = float(lat_raw) if lat_raw else None
                lng = float(lng_raw) if lng_raw else None

                prop = Property.objects.create(
                    workspace=request.workspace,
                    name=name,
                    property_type=prop_type,
                    description=row.get("description", ""),
                    estate_land_title=row.get("estate_land_title", ""),
                    status=status,
                )
                PropertyLocation.objects.create(
                    workspace=request.workspace,
                    property=prop,
                    address=row.get("address", "").strip() or f"{city}, {state}",
                    city=city,
                    state=state,
                    country="Nigeria",
                    postal_code=row.get("postal_code", "").strip() or None,
                    nearest_landmark=row.get("nearest_landmark", "").strip() or None,
                    latitude=lat,
                    longitude=lng,
                )
                created += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        return Response({
            "total_rows": len(rows),
            "created": created,
            "skipped": skipped,
            "error_count": len(errors),
            "errors": errors,
        })


# ---------------------------------------------------------------------------
# Template view
# ---------------------------------------------------------------------------

class PropertyBulkTemplateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        fmt = request.query_params.get("format", "csv").lower()
        if fmt == "xlsx":
            return self._xlsx_response()
        return self._csv_response()

    # ── CSV ──────────────────────────────────────────────────────────────────

    def _csv_response(self):
        buf = io.StringIO()
        writer = csv.writer(buf)

        # Header row
        writer.writerow(PROPERTY_TEMPLATE_HEADERS)

        # Description row (prefixed with # so the importer skips it)
        writer.writerow([
            f"# {'REQUIRED' if col[1] else 'optional'} — {col[2]}{(' Valid values: ' + col[3]) if col[3] else ''}"
            for col in PROPERTY_COLUMNS
        ])

        # Sample data rows
        for row in PROPERTY_SAMPLE_ROWS:
            writer.writerow(row)

        response = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="terratrail_properties_template.csv"'
        return response

    # ── XLSX ─────────────────────────────────────────────────────────────────

    def _xlsx_response(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: Data ────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Properties"

        BRAND_BLUE   = "0E2C72"
        BRAND_GREEN  = "10B981"
        REQUIRED_RED = "FEE2E2"
        OPT_GREY     = "F3F4F6"
        WHITE        = "FFFFFF"
        LIGHT_BLUE   = "EFF6FF"

        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Row 1: Column headers
        for col_idx, col in enumerate(PROPERTY_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col[0])
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Row 2: Field descriptions
        for col_idx, col in enumerate(PROPERTY_COLUMNS, start=1):
            req_label = "★ Required" if col[1] else "Optional"
            cell = ws.cell(row=2, column=col_idx, value=f"{req_label} — {col[2]}")
            cell.font = Font(italic=True, color="6B7280", size=8)
            cell.fill = PatternFill("solid", fgColor=REQUIRED_RED if col[1] else OPT_GREY)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 32

        # Sample data rows (rows 3+)
        for r_idx, sample in enumerate(PROPERTY_SAMPLE_ROWS, start=3):
            for c_idx, val in enumerate(sample, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if r_idx % 2 == 1 else WHITE)
                cell.font = Font(size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[r_idx].height = 18

        # Column widths
        col_widths = [28, 20, 50, 28, 12, 18, 14, 36, 14, 30, 12, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header rows
        ws.freeze_panes = "A3"

        # Data validation — property_type dropdown (column B = col 2)
        dv_type = DataValidation(
            type="list",
            formula1='"RESIDENTIAL_LAND,FARM_LAND,COMMERCIAL,MIXED_USE"',
            allow_blank=True,
            showDropDown=False,
        )
        dv_type.error = "Must be one of: RESIDENTIAL_LAND, FARM_LAND, COMMERCIAL, MIXED_USE"
        dv_type.errorTitle = "Invalid property type"
        dv_type.prompt = "Select from list"
        dv_type.promptTitle = "Property Type"
        ws.add_data_validation(dv_type)
        dv_type.add(f"B3:B502")

        # Data validation — status dropdown (column E = col 5)
        dv_status = DataValidation(
            type="list",
            formula1='"DRAFT,PUBLISHED"',
            allow_blank=True,
            showDropDown=False,
        )
        dv_status.error = "Must be DRAFT or PUBLISHED"
        dv_status.errorTitle = "Invalid status"
        ws.add_data_validation(dv_status)
        dv_status.add("E3:E502")

        # ── Sheet 2: Instructions ────────────────────────────────────────────
        wi = wb.create_sheet("Instructions")
        wi.sheet_view.showGridLines = False

        def iwrite(row, col, val, bold=False, size=11, color="1F2937", fill=None, wrap=False, indent=0):
            cell = wi.cell(row=row, column=col, value=(" " * indent) + str(val) if indent else val)
            cell.font = Font(bold=bold, size=size, color=color)
            cell.alignment = Alignment(wrap_text=wrap, vertical="top")
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            return cell

        wi.column_dimensions["A"].width = 3
        wi.column_dimensions["B"].width = 22
        wi.column_dimensions["C"].width = 10
        wi.column_dimensions["D"].width = 55
        wi.column_dimensions["E"].width = 35

        r = 1
        iwrite(r, 2, "Terratrail — Properties Bulk Upload Template", bold=True, size=14, color=BRAND_BLUE)
        r += 1
        iwrite(r, 2, "Fill in the 'Properties' sheet and upload it at Settings › Bulk Upload.")
        r += 2

        iwrite(r, 2, "HOW TO USE THIS TEMPLATE", bold=True, size=11, color=BRAND_BLUE)
        r += 1
        steps = [
            "1.  Do NOT rename or delete any column in the Properties sheet.",
            "2.  Do NOT remove row 1 (headers) or row 2 (descriptions).",
            "3.  Start entering your data from row 3 onwards.",
            "4.  Delete the sample rows (rows 3–7) before uploading, or leave them — the importer skips rows where 'name' already exists in your workspace.",
            "5.  Save as .xlsx and upload. Or export as CSV and upload.",
            "6.  Maximum 500 data rows per upload. Split larger files into batches.",
        ]
        for step in steps:
            iwrite(r, 2, step, wrap=True)
            wi.row_dimensions[r].height = 18
            r += 1

        r += 1
        iwrite(r, 2, "COLUMN REFERENCE", bold=True, size=11, color=BRAND_BLUE)
        r += 1

        # Table header
        for col_i, label in enumerate(["Column", "Required?", "Description & Valid Values"], start=2):
            c = wi.cell(row=r, column=col_i, value=label)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
            c.alignment = Alignment(horizontal="center")
        r += 1

        for col in PROPERTY_COLUMNS:
            col_name, required, desc, valid, example = col
            req_cell = wi.cell(row=r, column=3, value="★ Required" if required else "Optional")
            req_cell.font = Font(bold=required, color="DC2626" if required else "6B7280", size=9)
            req_cell.alignment = Alignment(horizontal="center")

            name_cell = wi.cell(row=r, column=2, value=col_name)
            name_cell.font = Font(bold=True, color=BRAND_BLUE, size=10)
            name_cell.fill = PatternFill("solid", fgColor=REQUIRED_RED if required else OPT_GREY)

            full_desc = desc
            if valid:
                full_desc += f"\n  Valid values: {valid.replace('|', ' · ')}"
            if example:
                full_desc += f"\n  Example: {example}"
            desc_cell = wi.cell(row=r, column=4, value=full_desc)
            desc_cell.font = Font(size=9)
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

            wi.row_dimensions[r].height = 42 if valid or example else 18
            r += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="terratrail_properties_template.xlsx"'
        wb.save(response)
        return response
