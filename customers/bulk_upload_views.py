"""
Bulk upload views for Customers (CSV / Excel).

POST /api/v1/customers/bulk-upload/
    Accepts a multipart file (CSV or .xlsx).
    Returns a summary: created, skipped (duplicate email), errors per row.

GET  /api/v1/customers/bulk-upload/template/?format=csv|xlsx
    Returns a downloadable template with sample data.
"""

import csv
import io

from django.http import HttpResponse
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import IsWorkspaceAdminOrReadOnly
from customers.models import Customer

# ---------------------------------------------------------------------------
# Column definitions — single source of truth for template + docs
# ---------------------------------------------------------------------------

CUSTOMER_COLUMNS = [
    # (header, required, description, valid_values_or_None, example)
    ("full_name",               True,  "Customer's full name (first + last)",                                              None,                                                         "Adebayo Johnson"),
    ("email",                   True,  "Email address — must be unique per workspace. Used to skip duplicates.",           None,                                                         "adebayo.johnson@email.com"),
    ("phone",                   False, "Phone number including country code",                                              None,                                                         "+2348012345678"),
    ("address",                 False, "Home or office address",                                                           None,                                                         "12 Broad Street, Lagos Island"),
    ("next_of_kin_name",        False, "Full name of next of kin",                                                         None,                                                         "Mrs. Funke Johnson"),
    ("next_of_kin_phone",       False, "Phone number of next of kin",                                                      None,                                                         "+2348098765432"),
    ("next_of_kin_relationship",False, "Relationship to the customer",                                                     None,                                                         "Mother"),
    ("referral_source",         False, "How the customer found you. Defaults to WALK_IN if blank.",                        "WALK_IN|REFERRAL|SOCIAL_MEDIA|WEBSITE|AGENT|OTHER",          "REFERRAL"),
]

CUSTOMER_TEMPLATE_HEADERS = [c[0] for c in CUSTOMER_COLUMNS]

# Multiple sample rows covering different referral sources and scenarios
CUSTOMER_SAMPLE_ROWS = [
    ["Adebayo Johnson",    "adebayo.johnson@email.com",    "+2348012345678", "12 Broad Street, Lagos Island",          "Mrs. Funke Johnson",   "+2348098765432", "Mother",  "REFERRAL"],
    ["Chidinma Okafor",    "chidinma.okafor@gmail.com",    "+2348123456789", "Plot 5 Galadimawa, Abuja",               "Mr. Emeka Okafor",     "+2347011223344", "Husband", "SOCIAL_MEDIA"],
    ["Oluwaseun Adeyemi",  "seun.adeyemi@yahoo.com",       "+2349012345678", "34 Allen Avenue, Ikeja, Lagos",          "",                     "",               "",        "WEBSITE"],
    ["Fatima Al-Hassan",   "fatima.alhassan@gmail.com",    "+2348034567890", "Maiduguri Road, Kano",                   "Alhaji Musa Al-Hassan","+2348045678901", "Father",  "AGENT"],
    ["Emeka Eze",          "emeka.eze@company.ng",         "+2347034567890", "3 Trans-Amadi, Port Harcourt",           "Mrs. Ngozi Eze",       "+2348056789012", "Wife",    "WALK_IN"],
    ["Aisha Bello",        "aisha.bello@outlook.com",      "+2348090123456", "Freedom Way, Lekki Phase 1, Lagos",      "Mr. Ibrahim Bello",    "+2348078901234", "Brother", "REFERRAL"],
    ["Tunde Ogundimu",     "tunde.ogundimu@hotmail.com",   "+2347056789012", "University Road, Ile-Ife, Osun",         "",                     "",               "",        "OTHER"],
    ["Blessing Nwachukwu", "blessing.nwachukwu@email.com", "+2348012398765", "15 Onitsha Road, Enugu",                 "Mr. Chukwu Nwachukwu", "+2348023456789", "Spouse",  "SOCIAL_MEDIA"],
]

REFERRAL_SOURCE_MAP = {
    "walk-in": "WALK_IN",
    "walk_in": "WALK_IN",
    "walkin":  "WALK_IN",
    "walk in": "WALK_IN",
    "referral": "REFERRAL",
    "social media": "SOCIAL_MEDIA",
    "social_media": "SOCIAL_MEDIA",
    "socialmedia":  "SOCIAL_MEDIA",
    "website": "WEBSITE",
    "web":     "WEBSITE",
    "agent":   "AGENT",
    "other":   "OTHER",
}
VALID_REFERRAL_SOURCES = {c[0] for c in Customer.ReferralSource.choices}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_rows(file_obj, filename: str):
    """Return list of row-dicts from CSV or xlsx."""
    name = filename.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl is required for Excel uploads. Use CSV instead.")
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb["Customers"] if "Customers" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        data_rows = []
        for row in rows[1:]:
            d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            if all(not v for v in d.values()):
                continue
            data_rows.append(d)
        return data_rows
    else:
        text = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            d = {k.strip().lower(): (v.strip() if v else "") for k, v in row.items()}
            if d.get(CUSTOMER_TEMPLATE_HEADERS[0], "").startswith("#"):
                continue
            rows.append(d)
        return rows


# ---------------------------------------------------------------------------
# Upload view
# ---------------------------------------------------------------------------

class CustomerBulkUploadView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated, IsWorkspaceAdminOrReadOnly]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            rows = _parse_rows(file, file.name)
        except Exception as e:
            return Response({"detail": str(e)}, status=400)

        if not rows:
            return Response({"detail": "The file contains no data rows."}, status=400)

        if len(rows) > 500:
            return Response({"detail": "Maximum 500 rows per upload. Please split your file."}, status=400)

        created = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(rows, start=2):
            full_name = row.get("full_name", "").strip()
            email     = row.get("email", "").strip().lower()

            if not full_name:
                errors.append({"row": idx, "error": "full_name is required"})
                continue
            if not email:
                errors.append({"row": idx, "error": "email is required"})
                continue
            if "@" not in email or "." not in email.split("@")[-1]:
                errors.append({"row": idx, "error": f"invalid email address: '{email}'"})
                continue

            if Customer.objects.filter(workspace=request.workspace, email__iexact=email).exists():
                skipped += 1
                continue

            raw_ref = row.get("referral_source", "").strip().lower()
            ref_source = (
                REFERRAL_SOURCE_MAP.get(raw_ref)
                or (raw_ref.upper() if raw_ref.upper() in VALID_REFERRAL_SOURCES else "WALK_IN")
            )

            try:
                Customer.objects.create(
                    workspace=request.workspace,
                    full_name=full_name,
                    email=email,
                    phone=row.get("phone", "").strip(),
                    address=row.get("address", "").strip(),
                    next_of_kin_name=row.get("next_of_kin_name", "").strip(),
                    next_of_kin_phone=row.get("next_of_kin_phone", "").strip(),
                    next_of_kin_relationship=row.get("next_of_kin_relationship", "").strip(),
                    referral_source=ref_source,
                )
                created += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        return Response({
            "total_rows": len(rows),
            "created": created,
            "skipped": skipped,
            "error_count": len(errors),
            "errors": errors,
        })


# ---------------------------------------------------------------------------
# Template view
# ---------------------------------------------------------------------------

class CustomerBulkTemplateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        fmt = request.query_params.get("format", "csv").lower()
        if fmt == "xlsx":
            return self._xlsx_response()
        return self._csv_response()

    # ── CSV ──────────────────────────────────────────────────────────────────

    def _csv_response(self):
        buf = io.StringIO()
        writer = csv.writer(buf)

        writer.writerow(CUSTOMER_TEMPLATE_HEADERS)

        # Description row (prefixed # so importer skips it)
        writer.writerow([
            f"# {'REQUIRED' if col[1] else 'optional'} — {col[2]}{(' Valid values: ' + col[3]) if col[3] else ''}"
            for col in CUSTOMER_COLUMNS
        ])

        for row in CUSTOMER_SAMPLE_ROWS:
            writer.writerow(row)

        response = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="terratrail_customers_template.csv"'
        return response

    # ── XLSX ─────────────────────────────────────────────────────────────────

    def _xlsx_response(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: Data ────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Customers"

        BRAND_BLUE   = "0E2C72"
        REQUIRED_RED = "FEE2E2"
        OPT_GREY     = "F3F4F6"
        WHITE        = "FFFFFF"
        LIGHT_BLUE   = "EFF6FF"

        thin   = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Row 1: Column headers
        for col_idx, col in enumerate(CUSTOMER_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col[0])
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Row 2: Field descriptions
        for col_idx, col in enumerate(CUSTOMER_COLUMNS, start=1):
            req_label = "★ Required" if col[1] else "Optional"
            cell = ws.cell(row=2, column=col_idx, value=f"{req_label} — {col[2]}")
            cell.font = Font(italic=True, color="6B7280", size=8)
            cell.fill = PatternFill("solid", fgColor=REQUIRED_RED if col[1] else OPT_GREY)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 32

        # Sample data rows (rows 3+)
        for r_idx, sample in enumerate(CUSTOMER_SAMPLE_ROWS, start=3):
            for c_idx, val in enumerate(sample, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if r_idx % 2 == 1 else WHITE)
                cell.font = Font(size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.border = border
            ws.row_dimensions[r_idx].height = 18

        # Column widths
        col_widths = [28, 34, 20, 38, 26, 20, 24, 18]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A3"

        # Data validation — referral_source dropdown (column H = col 8)
        dv_ref = DataValidation(
            type="list",
            formula1='"WALK_IN,REFERRAL,SOCIAL_MEDIA,WEBSITE,AGENT,OTHER"',
            allow_blank=True,
            showDropDown=False,
        )
        dv_ref.error = "Must be one of: WALK_IN, REFERRAL, SOCIAL_MEDIA, WEBSITE, AGENT, OTHER"
        dv_ref.errorTitle = "Invalid referral source"
        dv_ref.prompt = "Select from list"
        dv_ref.promptTitle = "Referral Source"
        ws.add_data_validation(dv_ref)
        dv_ref.add("H3:H502")

        # ── Sheet 2: Instructions ────────────────────────────────────────────
        wi = wb.create_sheet("Instructions")
        wi.sheet_view.showGridLines = False

        def iwrite(row, col, val, bold=False, size=11, color="1F2937", fill=None, wrap=False):
            cell = wi.cell(row=row, column=col, value=val)
            cell.font = Font(bold=bold, size=size, color=color)
            cell.alignment = Alignment(wrap_text=wrap, vertical="top")
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            return cell

        wi.column_dimensions["A"].width = 3
        wi.column_dimensions["B"].width = 26
        wi.column_dimensions["C"].width = 10
        wi.column_dimensions["D"].width = 55
        wi.column_dimensions["E"].width = 35

        r = 1
        iwrite(r, 2, "Terratrail — Customers Bulk Upload Template", bold=True, size=14, color=BRAND_BLUE)
        r += 1
        iwrite(r, 2, "Fill in the 'Customers' sheet and upload it at Settings › Bulk Upload.")
        r += 2

        iwrite(r, 2, "HOW TO USE THIS TEMPLATE", bold=True, size=11, color=BRAND_BLUE)
        r += 1
        steps = [
            "1.  Do NOT rename or delete any column in the Customers sheet.",
            "2.  Do NOT remove row 1 (headers) or row 2 (descriptions).",
            "3.  Start entering your customer data from row 3 onwards.",
            "4.  The sample rows (3–10) can be deleted before uploading, or left — the importer skips rows where the email already exists.",
            "5.  Each email must be unique. Duplicate emails within this workspace are skipped, not overwritten.",
            "6.  Maximum 500 data rows per upload. Split larger files into batches.",
            "7.  Save as .xlsx and upload, or export as CSV.",
        ]
        for step in steps:
            iwrite(r, 2, step, wrap=True)
            wi.row_dimensions[r].height = 18
            r += 1

        r += 1
        iwrite(r, 2, "REFERRAL SOURCE VALUES", bold=True, size=11, color=BRAND_BLUE)
        r += 1
        ref_notes = [
            ("WALK_IN",      "Customer walked into your office directly"),
            ("REFERRAL",     "Referred by an existing customer or contact"),
            ("SOCIAL_MEDIA", "Found you via Instagram, Facebook, Twitter, etc."),
            ("WEBSITE",      "Found you via your company website"),
            ("AGENT",        "Brought in by a sales or marketing agent"),
            ("OTHER",        "Any other source not listed above"),
        ]
        for val, desc in ref_notes:
            val_cell = wi.cell(row=r, column=2, value=val)
            val_cell.font = Font(bold=True, color=BRAND_BLUE, size=10)
            val_cell.fill = PatternFill("solid", fgColor=OPT_GREY)
            desc_cell = wi.cell(row=r, column=3, value=desc)
            desc_cell.font = Font(size=9)
            wi.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            wi.row_dimensions[r].height = 16
            r += 1

        r += 1
        iwrite(r, 2, "COLUMN REFERENCE", bold=True, size=11, color=BRAND_BLUE)
        r += 1

        for col_i, label in enumerate(["Column", "Required?", "Description & Valid Values"], start=2):
            c = wi.cell(row=r, column=col_i, value=label)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
            c.alignment = Alignment(horizontal="center")
        r += 1

        for col in CUSTOMER_COLUMNS:
            col_name, required, desc, valid, example = col

            req_cell = wi.cell(row=r, column=3, value="★ Required" if required else "Optional")
            req_cell.font = Font(bold=required, color="DC2626" if required else "6B7280", size=9)
            req_cell.alignment = Alignment(horizontal="center")

            name_cell = wi.cell(row=r, column=2, value=col_name)
            name_cell.font = Font(bold=True, color=BRAND_BLUE, size=10)
            name_cell.fill = PatternFill("solid", fgColor=REQUIRED_RED if required else OPT_GREY)

            full_desc = desc
            if valid:
                full_desc += f"\n  Valid values: {valid.replace('|', ' · ')}"
            if example:
                full_desc += f"\n  Example: {example}"
            desc_cell = wi.cell(row=r, column=4, value=full_desc)
            desc_cell.font = Font(size=9)
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

            wi.row_dimensions[r].height = 42 if valid or example else 18
            r += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="terratrail_customers_template.xlsx"'
        wb.save(response)
        return response
